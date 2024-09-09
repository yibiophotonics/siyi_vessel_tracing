# -*- coding: utf-8 -*-
"""
Created on Fri Aug 16 15:40:09 2024

@author: Administrator
"""

from skimage import io
import numpy as np
import os
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
from PIL import Image
from scipy.stats import mode
from sklearn.metrics import confusion_matrix as mtrx
import shutil

#%% dictionary - subject no: vessel no: file address - new dataset resized
img_folder = "D:/siyi_vessle/ji_amir_resize/"
subjects = {}

for subject in os.listdir(img_folder):
    if len(subject)==6:
        for subject_files in os.listdir("".join((img_folder,subject,"/"))):
            if len(subject_files)<3:
                subject_no = int(subject_files)
                print(subject_no,end=": ")
                img_files = []
                for img_file in os.listdir("".join((img_folder,subject,"/",subject_files))):
                    if len(img_file) > 10 and img_file.endswith(".tif"):
                        img_files.append(img_file)
                vessels = {}
                for img_file in img_files:
                    vessel_no = img_file.split("_")[-1]
                    vessel_no = vessel_no.split(".")[0]
                    vessel_no = int(vessel_no)
                    print(vessel_no,end=", ")
                    
                    vessels[vessel_no] = "".join((img_folder,subject,"/",subject_files,"/",img_file))
                subjects[subject_no] = vessels
                print("end")
                
#%% dictionary - subject no: vessel no: file address - original dataset
img_folder = "C:/Users/Administrator/OneDrive - Johns Hopkins/2023 Siyi Chen vessel tracing/whole_image/mask/ground_truth/"
subjects = {}

for subject in os.listdir(img_folder):
    subject_no = int(subject)
    print(subject_no,end=": ")
    vessels = {}
    for vessel in os.listdir("".join((img_folder,subject,"/"))):
        if vessel.endswith(".png"):
            vessel_no = vessel.split("_")[2]
            vessel_no = int(vessel_no)
            print(vessel_no,end=", ")
            
            vessels[vessel_no] = "".join((img_folder,subject,"/",vessel))
    subjects[subject_no] = vessels
    print("end")

#%% match prediction to ground truth vessels via start point
start_pts = pd.read_excel("D:/siyi_vessle/ji_amir_resize/whole_image/summary_3.xlsx")
#start_pts = pd.read_excel("C:/Users/Administrator/OneDrive - Johns Hopkins/2023 Siyi Chen vessel tracing/whole_image/input/summary.xlsx")
subject_list = np.unique(start_pts["Subject"])

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet['A1'] = 'Subject'
sheet['B1'] = 'StartPoint'
sheet['C1'] = 'GroundTruth'

#%%
subject_match = {}
for subject in list(subjects.keys()):
    print(subject,end=" | ")
    vessels = subjects[subject]
    vessel_match = {}
    for subject_idx in start_pts.index[start_pts["Subject"]==subject]:
        vessel_no = start_pts["Vessel"][subject_idx]
        xsp = int(start_pts["Row_sp"][subject_idx])
        ysp = int(start_pts["Column_sp"][subject_idx])
        xop = int(start_pts["Row_other"][subject_idx])
        yop = int(start_pts["Column_other"][subject_idx])
        
        for key,value in vessels.items():
            mask = io.imread(value)
            if mask[xsp,ysp]!=0:
                vessel_match[vessel_no]=key
                print(vessel_no,end=": ")
                print(key,end=" | ")
                sheet.append([subject,vessel_no,key])
                break
            elif mask[xop,yop]!=0:
                vessel_match[vessel_no]=key
                print(vessel_no,end=": ")
                print(key,end=" | ")
                sheet.append([subject,vessel_no,key])
                break
        if vessel_no not in vessel_match:
            print(vessel_no,end=": (")
            print(xsp,ysp,end=") | ")
            sheet.append([subject,vessel_no,'NONE'])
    subject_match[subject] = vessel_match
    print("end")
    
#workbook.save("D:/siyi_vessle/ji_amir_resize/whole_image/vessel_match_3.xlsx")

#%% match prediction to ground truth vessels via start point - multiple start point
start_pts = pd.read_excel("C:/Users/Administrator/OneDrive - Johns Hopkins/2023 Siyi Chen vessel tracing/whole_image/input/multiple start point/summary.xlsx")

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet['A1'] = 'Subject'
sheet['B1'] = 'Vessel'
sheet['C1'] = 'Sample'
sheet['D1'] = 'GroundTruth'

subject_match = {}
for subject in list(subjects.keys()):
    print(subject,end=" | ")
    vessels = subjects[subject]
    vessel_match = {}
    for subject_idx in start_pts.index[start_pts["Subject"]==subject]:
        vessel_no = start_pts["Vessel"][subject_idx]
        sample_no = start_pts["Sample"][subject_idx]
        
        xsp = int(start_pts["Row_sp"][subject_idx])
        ysp = int(start_pts["Column_sp"][subject_idx])
        xop = int(start_pts["Row_other"][subject_idx])
        yop = int(start_pts["Column_other"][subject_idx])
        
        for key,value in vessels.items():
            mask = io.imread(value)[:,:,0]
            try:
                if mask[xsp,ysp]==0:
                    vessel_match[vessel_no]=key
                    print(vessel_no,end=": ")
                    print(key,end=" | ")
                    sheet.append([subject,vessel_no,sample_no,key])
                    break
                elif mask[xop,yop]==0:
                    vessel_match[vessel_no]=key
                    print(vessel_no,end=": ")
                    print(key,end=" | ")
                    sheet.append([subject,vessel_no,sample_no,key])
                    break
            except:
                continue
        if vessel_no not in vessel_match:
            print(vessel_no,end=": (")
            print(xsp,ysp,end=") | ")
            sheet.append([subject,vessel_no,sample_no,'NONE'])
    subject_match[subject] = vessel_match
    print("end")

workbook.save("C:/Users/Administrator/OneDrive - Johns Hopkins/2023 Siyi Chen vessel tracing/whole_image/input/multiple start point/vessel_match.xlsx")
workbook.close()

#%% multiple start point - copy ground truth and match trace name
tracing = pd.read_excel("C:/Users/Administrator/OneDrive - Johns Hopkins/2023 Siyi Chen vessel tracing/whole_image/input/multiple start point/vessel_match.xlsx")
gt_path = "C:/Users/Administrator/OneDrive - Johns Hopkins/2023 Siyi Chen vessel tracing/whole_image/input/multiple start point/ground_truth/"

for subject in list(subjects.keys()):
    print(subject,end=" | ")
    gt_dir = "".join((gt_path,str(subject)))
    if not os.path.exists(gt_dir):
        os.makedirs(gt_dir)
    
    vessels = subjects[subject]
    
    subject_idx = tracing.loc[tracing["Subject"]==subject]
    vessel = np.unique(subject_idx["Vessel"])
    for v in vessel:
        vessel_idx = subject_idx.loc[subject_idx["Vessel"]==v]
        true_idx = mode(vessel_idx["GroundTruth"])[0]
        print(v,end=": ")
        print(true_idx,end=" | ")
        
        mask = vessels[true_idx]
        mask_name = "".join((gt_dir,"/",str(v),"_",str(true_idx),".png"))
        shutil.copy(mask, mask_name)
        
    print("end")


#%% plot start points in image
img_folder = "D:/siyi_vessle/ji_amir_resize/"
start_pts = pd.read_excel("D:/siyi_vessle/ji_amir_resize/whole_image/summary.xlsx")

for subject in os.listdir(img_folder)[1:]:
    if len(subject)==6:
        for subject_files in os.listdir("".join((img_folder,subject,"/"))):
            if len(subject_files)<3:
                subject_no = int(subject_files)
                vessels = subjects[subject_no]
                img = io.imread("".join((img_folder,subject,"/",subject_files,"/",subject,".tif")))
                #plt.figure(num=0,figsize=(15,15),clear=True)
                #plt.imshow(img)
                
                for subject_idx in start_pts.index[start_pts["Subject"]==subject_no]:
                    vessel_no = start_pts["Vessel"][subject_idx]
                    xsp = int(start_pts["Row_sp"][subject_idx])
                    ysp = int(start_pts["Column_sp"][subject_idx])
                    xop = int(start_pts["Row_other"][subject_idx])
                    yop = int(start_pts["Column_other"][subject_idx])
                    
                    '''
                    plt.annotate(str(vessel_no), xy=(ysp,xsp), xytext=(yop,xop), fontsize='small', arrowprops=dict(width=1,headwidth=3,headlength=3))
                    plt.scatter([ysp,yop],[xsp,xop],color="white",s=1)
                    plt.axis('off')
                    plt.savefig("".join((img_folder,subject,"/",subject_files,"/check_points_2.png")),bbox_inches='tight',pad_inches=0)
                    '''
                    
                    black_dot = True
                    for key,value in vessels.items():
                        mask = io.imread(value)
                        if mask[xsp,ysp]>0:
                            img[xsp,ysp,:] = 0
                            black_dot = False
                            break
                    if black_dot:
                        img[xsp,ysp,:] = 255
                        print("".join((subject_files,": ",str(vessel_no)," start point")))
                    black_dot = True
                    for key,value in vessels.items():
                        mask = io.imread(value)
                        if mask[xop,yop]>0:
                            img[xop,yop,:] = 0
                            black_dot = False
                            break
                    if black_dot:
                        img[xop,yop,:] = 255
                        print("".join((subject_files,": ",str(vessel_no)," other point")))
                            
                img = Image.fromarray(img)
                img.save("".join((img_folder,subject,"/",subject_files,"/check_points_2.png")))
                
#%% compare prediction and ground truth - single start point
path = "D:/siyi_vessle/ji_amir_resize/whole_image/vessel_match/"
gt_path = "".join((path,"ground_truth/"))
pd_path = "".join((path,"prediction/"))

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet['A1'] = 'Subject'
sheet['B1'] = 'Vessel'
sheet['C1'] = 'TN'
sheet['D1'] = 'FP'
sheet['E1'] = 'FN'
sheet['F1'] = 'TP'

for subject in os.listdir(gt_path):
    sj_gt = "".join((gt_path,subject,"/"))
    sj_pd = "".join((pd_path,subject,"/"))
    
    for vessel in os.listdir(sj_gt):
        igt = io.imread("".join((sj_gt,vessel)))
        if igt.ndim > 2:
            igt = igt[:,:,0]
        igt = igt / igt.max()
        if mode(igt,axis=None)[0] == 1:
            igt = 1 - igt
        
        vpd = vessel.split(".")[0]
        ipd = io.imread("".join((sj_pd,vpd,".png")))
        if ipd.ndim > 2:
            ipd = ipd[:,:,0]
        ipd = ipd / ipd.max()
        if mode(ipd,axis=None)[0] == 1:
            ipd = 1 - ipd
            
        tn, fp, fn, tp = mtrx(igt.ravel(), ipd.ravel()).ravel()
        sheet.append([int(subject),int(vpd),tn,fp,fn,tp])

workbook.save("".join((path,"result_stats.xlsx")))

#%% compare prediction and ground truth - multiple start point
path = "C:/Users/Administrator/OneDrive - Johns Hopkins/2023 Siyi Chen vessel tracing/whole_image/input/multiple start point/"
gt_path = "".join((path,"ground_truth/"))
pd_path = "".join((path,"tracing/"))

# get subject - vessel key
subjects = {}
for subject in os.listdir(gt_path):
    subject_no = int(subject)
    print(subject_no,end=": ")
    vessels = {}
    for vessel in os.listdir("".join((gt_path,subject,"/"))):
        vessel_no = vessel.split("_")[0]
        vessel_no = int(vessel_no)
        print(vessel_no,end=", ")
            
        vessels[vessel_no] = "".join((gt_path,subject,"/",vessel))
    subjects[subject_no] = vessels
    print("end")

#%% stats
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet['A1'] = 'Subject'
sheet['B1'] = 'Vessel'
sheet['C1'] = 'TN'
sheet['D1'] = 'FP'
sheet['E1'] = 'FN'
sheet['F1'] = 'TP'
sheet['G1'] = 'Sample'
#%
for subject in list(subjects.keys()):
    sj_pd = "".join((pd_path,str(subject),"/map/"))
    
    for vessel in os.listdir(sj_pd):
        if not vessel.endswith("post.png"):
            continue
        vessel_no = int(vessel.split("_")[0])
        sample_no = int(vessel.split("_")[1])
        
        igt = io.imread(subjects[subject][vessel_no])
        if igt.ndim > 2:
            igt = igt[:,:,0]
        igt = igt / igt.max()
        if mode(igt,axis=None)[0] == 1:
            igt = 1 - igt
        
        ipd = io.imread("".join((sj_pd,vessel)))
        if ipd.ndim > 2:
            ipd = ipd[:,:,0]
        ipd = ipd / ipd.max()
        if mode(ipd,axis=None)[0] == 1:
            ipd = 1 - ipd
            
        tn, fp, fn, tp = mtrx(igt.ravel(), ipd.ravel()).ravel()
        sheet.append([subject,vessel_no,tn,fp,fn,tp,sample_no])

workbook.save("".join((path,"result_stats.xlsx")))